from odoo import http
from odoo.http import request
from odoo.exceptions import AccessError, UserError, ValidationError
from io import BytesIO
import base64
import re
import datetime
import csv
from io import StringIO
import logging




from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


class HrmisSanctionedPostsController(http.Controller):


    def _slug_part(self, s: str) -> str:
        """
        For login parts: keep letters+digits only, lowercase.
        """
        s = "" if s is None else str(s)
        s = s.strip().lower()
        s = re.sub(r"[^a-z0-9]+", "", s)
        return s

    def _strip_titles(self, name: str) -> str:
        """
        Removes common prefixes like Dr., Prof., etc.
        """
        name = "" if name is None else str(name)
        name = name.strip()

        # remove leading "Dr", "Dr.", "Doctor", etc (case-insensitive)
        name = re.sub(r"^(dr\.?|doctor|prof\.?|professor)\s+", "", name, flags=re.I)

        # optional: remove extra whitespace
        name = re.sub(r"\s+", " ", name).strip()
        return name

    def _first_last_from_doctor_name(self, full_name: str):
        """
        Extract firstname/lastname from "Name of doctor".
        Strategy:
        - Strip titles
        - Remove everything after relational markers like S/o, D/o, W/o (because it adds noise)
        - Split remaining by spaces; first token = first name, last token = last name
        """
        n = self._strip_titles(full_name)

        # cut off at S/o, D/o, W/o etc (case-insensitive)
        n = re.split(r"\b(s/o|d/o|w/o)\b", n, flags=re.I)[0].strip()

        # remove trailing punctuation
        n = re.sub(r"[.,]+$", "", n).strip()
        n = re.sub(r"\s+", " ", n).strip()

        parts = [p for p in n.split(" ") if p.strip()]
        if len(parts) < 2:
            return (parts[0] if parts else "", "")
        return (parts[0], parts[-1])

    def _dob_digits(self, dob_val) -> str:
        """
        New requirement: DOB digits only, keep natural DDMMYYYY if date object.
        If string => strip non-digits (04.02.1979 => 04021979).
        """
        if isinstance(dob_val, (datetime.date, datetime.datetime)):
            return dob_val.strftime("%d%m%Y")
        return self._digits_only("" if dob_val is None else str(dob_val))

    def _domicile_key(self, domicile_val) -> str:
        """
        last three letters of domicile (letters only), lowercase.
        If < 3 letters, use whatever exists.
        """
        s = "" if domicile_val is None else str(domicile_val)
        letters_only = re.sub(r"[^A-Za-z]+", "", s).lower()
        if not letters_only:
            return ""
        return letters_only[-3:] if len(letters_only) >= 3 else letters_only
    # ---- helpers ----
    def _ensure_admin(self):
        if not request.env.user.has_group("base.group_system"):
            raise AccessError("You are not allowed to access this page.")

    def _digits_only(self, s: str) -> str:
        return re.sub(r"\D+", "", s or "")

    def _last4_digits(self, s) -> str:
        d = self._digits_only("" if s is None else str(s))
        return d[-4:] if len(d) >= 4 else d.zfill(4)

    def _dob_digits(self, dob_val) -> str:
        """
        As per requirement: DOB without dashes/slashes.
        If excel date => format YYYYMMDD
        If string => strip non-digits (02-19-2021 => 02192021)
        """
        if isinstance(dob_val, (datetime.date, datetime.datetime)):
            return dob_val.strftime("%Y%m%d")
        return self._digits_only("" if dob_val is None else str(dob_val))

    def _normalize_columns(self, header_row_values):
        header = [("" if v is None else str(v).strip()) for v in header_row_values]
        columns, used = [], set()
        for i, h in enumerate(header, start=1):
            name = h if h else f"Column {i}"
            base = name
            n = 2
            while name in used:
                name = f"{base} ({n})"
                n += 1
            used.add(name)
            columns.append(name)
        return columns

    def _find_col(self, columns, candidates):
        """
        candidates: list of possible header names (case-insensitive)
        returns the actual column name as in columns[] or None
        """
        m = {c.lower(): c for c in columns}
        for cand in candidates:
            key = cand.lower()
            if key in m:
                return m[key]
        return None

    def _template(self):
        # Keep your existing external id (as you already used)
        return "sanctioned_posts.hrmis_sanctioned_posts_upload_page"

    # ---- routes ----
    @http.route("/hrmis/sanctioned_posts/upload", type="http", auth="user", website=True)
    def upload_page(self, **kw):
        self._ensure_admin()
        return request.render(self._template(), {
            "active_menu": "sanctioned_posts",
            "upload_result": None,
            "upload_error": None,
            "attachment_id": None,
        })

    @http.route("/hrmis/sanctioned_posts/upload_xlsx", type="http", auth="user",
                website=True, csrf=True, methods=["POST"])
    def upload_xlsx(self, **post):
        self._ensure_admin()

        if openpyxl is None:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "openpyxl is not installed on the server. Install it to read .xlsx files.",
                "attachment_id": None,
            })

        file_storage = request.httprequest.files.get("xlsx_file")
        if not file_storage or not file_storage.filename:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "No file selected. Please upload an .xlsx file.",
                "attachment_id": None,
            })

        if not file_storage.filename.lower().endswith(".xlsx"):
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "Invalid file type. Please upload an .xlsx file.",
                "attachment_id": None,
            })

        try:
            content = file_storage.read()

            # Store file so "Create Users" can reuse it without re-upload
            attachment = request.env["ir.attachment"].sudo().create({
                "name": file_storage.filename,
                "type": "binary",
                "datas": base64.b64encode(content),
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "res_model": "res.users",
            })

            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active

            columns = self._normalize_columns([c.value for c in ws[1]])

            rows = []
            count = 0
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if not row_cells or all((v is None or str(v).strip() == "") for v in row_cells):
                    continue
                count += 1

                if len(rows) < 10:
                    row_dict = {}
                    for idx, col_name in enumerate(columns):
                        val = row_cells[idx] if idx < len(row_cells) else ""
                        row_dict[col_name] = "" if val is None else val
                    rows.append(row_dict)

            upload_result = {
                "count": count,
                "columns": columns,
                "top10": rows,
            }

            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": upload_result,
                "upload_error": None,
                "attachment_id": attachment.id,
            })

        except Exception as e:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": f"Failed to read XLSX: {e}",
                "attachment_id": None,
            })

    
    @http.route(
        "/hrmis/sanctioned_posts/create_users",
        type="http",
        auth="user",
        website=True,
        csrf=True,
        methods=["POST"],
    )
    def create_users(self, **post):
        self._ensure_admin()

        attachment_id = int(post.get("attachment_id") or 0)
        if not attachment_id:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "Missing attachment_id. Please upload the XLSX again.",
                "attachment_id": None,
                "creds_attachment_id": None,
                "failed_attachment_id": None,
                "creds_preview": [],
            })

        attachment = request.env["ir.attachment"].sudo().browse(attachment_id)
        if not attachment.exists():
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "Uploaded file not found. Please upload again.",
                "attachment_id": None,
                "creds_attachment_id": None,
                "failed_attachment_id": None,
                "creds_preview": [],
            })

        if openpyxl is None:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "openpyxl is not installed on the server.",
                "attachment_id": attachment.id,
                "creds_attachment_id": None,
                "failed_attachment_id": None,
                "creds_preview": [],
            })

        Users = request.env["res.users"].sudo()
        Attach = request.env["ir.attachment"].sudo()

        try:
            content = base64.b64decode(attachment.datas or b"")
            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active

            columns = self._normalize_columns([c.value for c in ws[1]])

            # Required headers for doctor sheet
            name_col = self._find_col(columns, ["name of doctor", "name", "doctor name"])
            dob_col = self._find_col(columns, ["date of birth", "dob", "date_of_birth"])
            domicile_col = self._find_col(columns, ["domicile"])

            if not name_col or not dob_col or not domicile_col:
                raise UserError(
                    "Missing required columns. Required: Name of doctor, Date of birth, Domicile.\n"
                    f"Detected columns: {', '.join(columns)}"
                )

            # ---------------------------
            # PASS 1: Collect candidate logins (no DB writes)
            # ---------------------------
            candidate_logins = set()
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if not row_cells or all((v is None or str(v).strip() == "") for v in row_cells):
                    continue
                row = {columns[i]: (row_cells[i] if i < len(row_cells) else None) for i in range(len(columns))}
                full_name = ("" if row.get(name_col) is None else str(row.get(name_col)).strip())
                if not full_name:
                    continue

                first, last = self._first_last_from_doctor_name(full_name)
                first_s = self._slug_part(first)
                last_s = self._slug_part(last)
                if first_s and last_s:
                    candidate_logins.add(f"{first_s}.{last_s}")

            # Query existing logins only for those in the sheet (1 SQL query)
            existing_logins = set()
            if candidate_logins:
                existing_logins = set(
                    rec["login"] for rec in Users.search_read([("login", "in", list(candidate_logins))], ["login"])
                )

            # Reset worksheet iterator (reload workbook)
            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active

            # ---------------------------
            # Import state
            # ---------------------------
            created = 0
            skipped = 0
            processed = 0

            errors_preview = []
            errors_preview_limit = 50

            failed_rows = []  # full downloadable list

            created_creds_all = []  # (login, password)
            creds_preview = []
            preview_limit = 50

            seen_logins = set()  # duplicates within sheet

            batch = []
            batch_creds = []     # (login, password)
            batch_meta = []      # meta dict for failure XLSX
            batch_size = 2000

            def add_cred(login, password):
                created_creds_all.append((login, password))
                if len(creds_preview) < preview_limit:
                    creds_preview.append({"login": login, "password": password})

            def log_fail(row_index, full_name, login, dob_digits, domicile, reason):
                nonlocal skipped
                skipped += 1
                if len(errors_preview) < errors_preview_limit:
                    errors_preview.append(f"Row {row_index}: {reason}")
                failed_rows.append({
                    "row": row_index,
                    "name": full_name or "",
                    "login": login or "",
                    "dob_digits": dob_digits or "",
                    "domicile": domicile or "",
                    "reason": reason or "",
                })

            def flush_batch():
                nonlocal created, batch, batch_creds, batch_meta
                if not batch:
                    return
                try:
                    new_users = Users.create(batch)
                    for _u, (lg, pw) in zip(new_users, batch_creds):
                        add_cred(lg, pw)
                        existing_logins.add(lg)  # keep set current
                    created += len(batch)
                    request.env.cr.commit()
                except Exception:
                    request.env.cr.rollback()
                    # isolate failing rows
                    for vals, (lg, pw), meta in zip(batch, batch_creds, batch_meta):
                        try:
                            Users.create(vals)
                            add_cred(lg, pw)
                            existing_logins.add(lg)
                            created += 1
                        except Exception as e2:
                            log_fail(
                                meta.get("row_index"),
                                meta.get("full_name"),
                                lg,
                                meta.get("dob_digits"),
                                meta.get("domicile"),
                                f"Create failed: {str(e2)}"
                            )
                    request.env.cr.commit()
                finally:
                    batch = []
                    batch_creds = []
                    batch_meta = []

            # ---------------------------
            # PASS 2: Validate + Create
            # ---------------------------
            for row_index, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row_cells or all((v is None or str(v).strip() == "") for v in row_cells):
                    continue

                processed += 1

                row = {columns[i]: (row_cells[i] if i < len(row_cells) else None) for i in range(len(columns))}

                full_name = ("" if row.get(name_col) is None else str(row.get(name_col)).strip())
                domicile = ("" if row.get(domicile_col) is None else str(row.get(domicile_col)).strip())
                dob_val = row.get(dob_col)

                if not full_name:
                    log_fail(row_index, full_name, "", "", domicile, "Missing Name of doctor")
                    continue
                if not domicile:
                    log_fail(row_index, full_name, "", "", domicile, "Missing Domicile")
                    continue

                dob_digits = self._dob_digits(dob_val)
                if not dob_digits:
                    log_fail(row_index, full_name, "", "", domicile, "Missing/invalid Date of birth")
                    continue

                first, last = self._first_last_from_doctor_name(full_name)
                first_s = self._slug_part(first)
                last_s = self._slug_part(last)

                if not first_s or not last_s:
                    log_fail(row_index, full_name, "", dob_digits, domicile, "Cannot parse firstname/lastname from Name of doctor")
                    continue

                login = f"{first_s}.{last_s}"

                # duplicates inside sheet
                if login in seen_logins:
                    log_fail(row_index, full_name, login, dob_digits, domicile, "Duplicate login in sheet")
                    continue
                seen_logins.add(login)

                # duplicates in DB (fast set check)
                if login in existing_logins:
                    log_fail(row_index, full_name, login, dob_digits, domicile, "Login already exists in database")
                    continue

                domicile_key = self._domicile_key(domicile)
                if not domicile_key:
                    log_fail(row_index, full_name, login, dob_digits, domicile, "Invalid domicile for domicilekey (needs letters)")
                    continue

                password = f"{last_s}{dob_digits}{domicile_key}"

                vals = {
                    "name": full_name,
                    "login": login,
                    "password": password,
                    "hrmis_role": "employee",
                }
                meta = {
                    "row_index": row_index,
                    "full_name": full_name,
                    "dob_digits": dob_digits,
                    "domicile": domicile,
                }

                batch.append(vals)
                batch_creds.append((login, password))
                batch_meta.append(meta)

                if len(batch) >= batch_size:
                    flush_batch()

            flush_batch()

            # ---------------------------
            # Credentials CSV attachment
            # ---------------------------
            csv_buf = StringIO()
            writer = csv.writer(csv_buf)
            writer.writerow(["login", "password"])
            for lg, pw in created_creds_all:
                writer.writerow([lg, pw])

            creds_attachment = Attach.create({
                "name": "created_users_credentials.csv",
                "type": "binary",
                "datas": base64.b64encode(csv_buf.getvalue().encode("utf-8")),
                "mimetype": "text/csv",
                "res_model": "res.users",
            })

            # ---------------------------
            # Failed XLSX attachment (ALL skipped rows)
            # ---------------------------
            failed_attachment = None
            if failed_rows:
                fail_wb = openpyxl.Workbook()
                fail_ws = fail_wb.active
                fail_ws.title = "Failed Users"

                headers = ["Row", "Name of doctor", "Login", "DOB Digits", "Domicile", "Reason"]
                fail_ws.append(headers)

                header_font = Font(bold=True)
                for col in range(1, len(headers) + 1):
                    c = fail_ws.cell(row=1, column=col)
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for fr in failed_rows:
                    fail_ws.append([
                        fr.get("row", ""),
                        fr.get("name", ""),
                        fr.get("login", ""),
                        fr.get("dob_digits", ""),
                        fr.get("domicile", ""),
                        fr.get("reason", ""),
                    ])

                fail_ws.freeze_panes = "A2"
                fail_ws.column_dimensions["A"].width = 8
                fail_ws.column_dimensions["B"].width = 45
                fail_ws.column_dimensions["C"].width = 25
                fail_ws.column_dimensions["D"].width = 14
                fail_ws.column_dimensions["E"].width = 20
                fail_ws.column_dimensions["F"].width = 55

                out = BytesIO()
                fail_wb.save(out)
                out.seek(0)

                failed_attachment = Attach.create({
                    "name": "failed_users.xlsx",
                    "type": "binary",
                    "datas": base64.b64encode(out.getvalue()),
                    "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "res_model": "res.users",
                })

            upload_result = {
                "count": processed,
                "columns": columns,
                "top10": [],
                "created": created,
                "skipped": skipped,
                "errors": errors_preview,
            }

            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": upload_result,
                "upload_error": None,
                "attachment_id": attachment.id,
                "creds_attachment_id": creds_attachment.id,
                "failed_attachment_id": failed_attachment.id if failed_attachment else None,
                "creds_preview": creds_preview,
            })

        except Exception as e:
            return request.render(self._template(), {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": f"Import failed: {e}",
                "attachment_id": attachment.id,
                "creds_attachment_id": None,
                "failed_attachment_id": None,
                "creds_preview": [],
            })

    @http.route(
        "/hrmis/sanctioned_posts/download_failed",
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
        website=True,
    )
    def download_failed(self, attachment_id=None, **kw):
        self._ensure_admin()

        try:
            att_id = int(attachment_id or 0)
        except Exception:
            att_id = 0

        if not att_id:
            raise UserError("Missing attachment_id for failed file.")

        attachment = request.env["ir.attachment"].sudo().browse(att_id).exists()
        if not attachment:
            raise UserError("Failed file not found.")

        content = base64.b64decode(attachment.datas or b"")
        filename = attachment.name or "failed_users.xlsx"

        return request.make_response(
            content,
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
            ],
        )

    # (optional) keep /hrmis/sanctioned_posts route if you want, but ensure template exists
    @http.route('/hrmis/sanctioned_posts', type='http', auth='user', website=True)
    def sanctioned_posts(self, **kw):
        self._ensure_admin()
        # If you don't have hr_holidays_updates.sanctioned_posts_template, you can redirect to upload:
        return request.redirect("/hrmis/sanctioned_posts/upload")

    @http.route(
        "/hrmis/sanctioned_posts/download_xlsx",
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
        website=True,
    )
    def download_xlsx(self, attachment_id=None, **kw):
        self._ensure_admin()

        try:
            att_id = int(attachment_id or 0)
        except Exception:
            att_id = 0

        if not att_id:
            raise UserError("Missing attachment_id. Please upload the XLSX again.")

        attachment = request.env["ir.attachment"].sudo().browse(att_id).exists()
        if not attachment:
            raise UserError("File not found. Please upload the XLSX again.")

        # (Optional) safety: ensure it looks like xlsx
        mimetype = attachment.mimetype or ""
        if "spreadsheetml" not in mimetype and not (attachment.name or "").lower().endswith(".xlsx"):
            _logger.warning("[SANCTIONED_POSTS] Attachment %s is not xlsx: name=%s mimetype=%s",
                            attachment.id, attachment.name, mimetype)

        _logger.info("[SANCTIONED_POSTS] Downloading uploaded XLSX attachment_id=%s name=%s",
                     attachment.id, attachment.name)

        content = base64.b64decode(attachment.datas or b"")
        filename = attachment.name or "sanctioned_posts.xlsx"

        return request.make_response(
            content,
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
            ],
        )


    @http.route(
        ["/hrmis/staff/download_details"],
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
        website=True,
    )
    def hrmis_staff_download_details(self, ids="all", **kw):
        """
        Export Staff Details XLSX (ONLY selected fields).

        - Main sheet: selected hr.employee fields (listed by you)
        - Separate sheets: service history, trainings, posting, promotion, qualification, leave
        - Prefer latest SUBMITTED hrmis.employee.profile.request values for overlapping fields
        - Exclude employees whose user_id.hrmis_role == 'section_officer'
        - ids=all OR ids=1,2,3
        """
        user = request.env.user
        _logger.info("[STAFF_EXPORT] /hrmis/staff/download_details ids=%s by %s(%s)", ids, user.name, user.id)

        # Access guard
        if not (
            user.has_group("hr.group_hr_user")
            or user.has_group("hr.group_hr_manager")
            or user.has_group("base.group_system")
        ):
            _logger.warning("[STAFF_EXPORT] Access denied user=%s(%s)", user.name, user.id)
            raise AccessError("You are not allowed to export staff details.")

        Employee = request.env["hr.employee"].with_context(active_test=False).sudo()
        ProfileReq = request.env["hrmis.employee.profile.request"].sudo()

        # Resolve employees
        if (ids or "all") == "all":
            employees = Employee.search([])
        else:
            try:
                emp_ids = [int(x) for x in (ids or "").split(",") if x.strip()]
            except Exception:
                _logger.exception("[STAFF_EXPORT] Invalid ids param: %s", ids)
                emp_ids = []
            employees = Employee.browse(emp_ids).exists()

        # Filter: exclude section_officer role
        employees = employees.filtered(lambda e: (not e.user_id) or (e.user_id.hrmis_role != "section_officer"))
        _logger.info("[STAFF_EXPORT] Employees after role filter: %s", len(employees))

        # ---------- formatting helpers ----------
        def fmt_date(d):
            return d.strftime("%Y-%m-%d") if d else ""

        def fmt_m2o(v):
            return v.display_name if v else ""

        def fmt_recordset_names(rs):
            if not rs:
                return ""
            try:
                return ", ".join(rs.mapped("display_name"))
            except Exception:
                return ", ".join(map(str, rs.ids))

        def val_from(emp, req, field_name):
            """
            Prefer submitted request value if field exists on request and is non-empty,
            otherwise use employee value.
            """
            emp_val = getattr(emp, field_name, False)
            if req and field_name in req._fields:
                req_val = getattr(req, field_name, False)
                if req_val not in (False, None, ""):
                    return req_val
            return emp_val

        # ---------- workbook ----------
        wb = Workbook()

        # MAIN SHEET
        ws = wb.active
        ws.title = "Staff"

        main_fields = [
            # identify
            ("name", "Employee Name"),
            ("work_email", "Work Email"),
            ("user_login", "User Login"),
            ("user_hrmis_role", "User HRMIS Role"),

            # your listed fields
            ("hrmis_employee_id", "Employee ID / Service Number"),
            ("hrmis_cnic", "CNIC"),
            ("birthday", "Date of Birth"),
            ("hrmis_commission_date", "Commission Date"),
            ("hrmis_father_name", "Father Name"),
            ("hrmis_joining_date", "Joining Date"),
            ("gender", "Gender"),
            ("hrmis_cadre", "Cadre"),
            ("hrmis_designation", "Designation"),
            ("hrmis_bps", "BPS Grade"),
            ("hrmis_merit_number", "Merit Number"),
            ("district_id", "Current District"),
            ("facility_id", "Current Facility"),
            ("hrmis_contact_info", "Contact Info"),
            ("hrmis_leaves_taken", "Total Leaves Taken (Days)"),

            # attachments filenames only (not binary)
            ("hrmis_cnic_front_filename", "CNIC Front Filename"),
            ("hrmis_cnic_back_filename", "CNIC Back Filename"),

            ("hrmis_domicile", "Domicile"),

            # qualification/promotions
            ("qualification", "Last Qualification Received"),
            ("qualification_date", "Qualification Date"),
            ("year_qualification", "Year of Qualification"),
            ("date_promotion", "Last Promotion Date"),
        ]

        headers = [h for _, h in main_fields]
        ws.append(headers)

        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_idx = 1

        for emp in employees:
            submitted_req = ProfileReq.search(
                [("employee_id", "=", emp.id), ("state", "=", "submitted")],
                order="id desc",
                limit=1,
            )

            row = []
            for fname, _label in main_fields:
                # synthetic columns
                if fname == "user_login":
                    row.append(emp.user_id.login if emp.user_id else "")
                    continue
                if fname == "user_hrmis_role":
                    row.append(emp.user_id.hrmis_role if emp.user_id else "")
                    continue

                v = val_from(emp, submitted_req, fname)

                # format by type
                fld = emp._fields.get(fname)
                if fld:
                    if fld.type == "many2one":
                        row.append(fmt_m2o(v))
                    elif fld.type in ("date", "datetime"):
                        # v may be date/datetime
                        row.append(fmt_date(v))
                    elif fld.type in ("one2many", "many2many"):
                        # keep main sheet clean: put counts only, details in separate sheets
                        row.append(len(v) if v else 0)
                    elif fld.type == "binary":
                        # don't dump binary, show YES/NO
                        row.append("YES" if v else "")
                    else:
                        row.append(v if v not in (None, False) else "")
                else:
                    row.append(v if v not in (None, False) else "")

            ws.append(row)
            row_idx += 1

        ws.freeze_panes = "A2"
        for col_idx in range(1, len(headers) + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max(12, len(headers[col_idx - 1]) + 2), 45)

        # ---------- Helper to create history sheets ----------
        def add_sheet(title, sheet_headers):
            sh = wb.create_sheet(title=title)
            sh.append(sheet_headers)
            for c in range(1, len(sheet_headers) + 1):
                cell = sh.cell(row=1, column=c)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sh.freeze_panes = "A2"
            return sh

        # SERVICE HISTORY (hrmis_service_history_ids)
        sh_service = add_sheet("Service History", [
            "Employee ID", "Employee Name",
            "District", "Facility",
            "From Date", "End Date",
            "Commission Date",
        ])
        for emp in employees:
            for rec in emp.hrmis_service_history_ids:
                sh_service.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    fmt_m2o(getattr(rec, "district_id", False)),
                    fmt_m2o(getattr(rec, "facility_id", False)),
                    fmt_date(getattr(rec, "from_date", False)),
                    fmt_date(getattr(rec, "end_date", False)),
                    fmt_date(getattr(rec, "commission_date", False)),
                ])

        # TRAININGS (hrmis_training_ids)
        sh_training = add_sheet("Trainings", [
            "Employee ID", "Employee Name",
            "Title", "Institute", "From Date", "End Date", "Remarks",
        ])
        for emp in employees:
            for rec in emp.hrmis_training_ids:
                sh_training.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    getattr(rec, "name", "") or getattr(rec, "title", "") or "",
                    getattr(rec, "institute", "") or "",
                    fmt_date(getattr(rec, "start_date", False) or getattr(rec, "from_date", False)),
                    fmt_date(getattr(rec, "end_date", False) or getattr(rec, "to_date", False)),
                    getattr(rec, "remarks", "") or "",
                ])

        # QUALIFICATION HISTORY (qualification_history_ids)
        sh_qh = add_sheet("Qualification History", [
            "Employee ID", "Employee Name",
            "Degree", "Specialization",
            "Start Date", "End Date",
        ])
        for emp in employees:
            for rec in emp.qualification_history_ids:
                sh_qh.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    getattr(rec, "degree", "") or "",
                    getattr(rec, "specialization", "") or "",
                    fmt_date(getattr(rec, "start_date", False)),
                    fmt_date(getattr(rec, "end_date", False)),
                ])

        # POSTING HISTORY (posting_history_ids)
        sh_posting = add_sheet("Posting History", [
            "Employee ID", "Employee Name",
            "District", "Facility", "Designation", "BPS",
            "Start Date", "End Date", "Is Current",
        ])
        for emp in employees:
            for rec in emp.posting_history_ids:
                sh_posting.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    fmt_m2o(getattr(rec, "district_id", False)),
                    fmt_m2o(getattr(rec, "facility_id", False)),
                    fmt_m2o(getattr(rec, "designation_id", False)),
                    getattr(rec, "bps", "") or "",
                    fmt_date(getattr(rec, "start_date", False)),
                    fmt_date(getattr(rec, "end_date", False)),
                    "YES" if getattr(rec, "is_current", False) else "",
                ])

        # PROMOTION HISTORY (promotion_history_ids)
        sh_promo = add_sheet("Promotion History", [
            "Employee ID", "Employee Name",
            "BPS From", "BPS To", "Promotion Date",
        ])
        for emp in employees:
            for rec in emp.promotion_history_ids:
                sh_promo.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    getattr(rec, "bps_from", "") or "",
                    getattr(rec, "bps_to", "") or "",
                    fmt_date(getattr(rec, "promotion_date", False)),
                ])

        # LEAVE HISTORY (leave_history_ids)
        sh_leave = add_sheet("Leave History", [
            "Employee ID", "Employee Name",
            "Leave Type", "Start Date", "End Date",
        ])
        for emp in employees:
            for rec in emp.leave_history_ids:
                sh_leave.append([
                    emp.hrmis_employee_id or "",
                    emp.name or "",
                    fmt_m2o(getattr(rec, "leave_type_id", False)),
                    fmt_date(getattr(rec, "start_date", False)),
                    fmt_date(getattr(rec, "end_date", False)),
                ])

        # Output
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"staff_details_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        _logger.info("[STAFF_EXPORT] Sending file %s", filename)

        return request.make_response(
            bio.getvalue(),
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
            ],
        )