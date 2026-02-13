from odoo import http
from odoo.http import request
from odoo.exceptions import AccessError, UserError
from io import BytesIO
import re
import datetime

try:
    import openpyxl
except Exception:
    openpyxl = None


class HrmisUserBulkImportController(http.Controller):

    def _ensure_admin(self):
        if not request.env.user.has_group("base.group_system"):
            raise AccessError("You are not allowed to access this page.")

    # ---------- helpers ----------
    def _digits_only(self, s: str) -> str:
        return re.sub(r"\D+", "", s or "")

    def _last4(self, s: str) -> str:
        d = self._digits_only(s)
        return d[-4:] if len(d) >= 4 else d.zfill(4)

    def _parse_dob_to_digits(self, dob_val) -> str:
        """
        Return DOB as digits only, as requested: dob without dashes/slashes.
        If date object => YYYYMMDD.
        If string => strip non-digits (e.g., 02-19-2021 => 02192021).
        """
        if isinstance(dob_val, (datetime.date, datetime.datetime)):
            return dob_val.strftime("%Y%m%d")
        return self._digits_only(str(dob_val or ""))

    def _get_cell_str(self, v) -> str:
        return "" if v is None else str(v).strip()

    def _normalize_columns(self, header_cells):
        header = [self._get_cell_str(c) for c in header_cells]
        columns, used = [], set()
        for i, h in enumerate(header, start=1):
            name = h or f"Column {i}"
            base = name
            n = 2
            while name in used:
                name = f"{base} ({n})"
                n += 1
            used.add(name)
            columns.append(name)
        return columns

    def _find_col(self, columns, candidates):
        """
        candidates: list of acceptable header names (case-insensitive)
        """
        lower_map = {c.lower(): c for c in columns}
        for cand in candidates:
            key = cand.lower()
            if key in lower_map:
                return lower_map[key]
        return None


    @http.route("/hrmis/users/import_preview", type="http", auth="user", website=True, csrf=True, methods=["POST"])
    def import_preview(self, **post):
        self._ensure_admin()

        if openpyxl is None:
            return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "openpyxl is not installed on the server.",
                "attachment_id": None,
            })

        fs = request.httprequest.files.get("xlsx_file")
        if not fs or not fs.filename:
            return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "No file selected.",
                "attachment_id": None,
            })

        if not fs.filename.lower().endswith(".xlsx"):
            return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": "Invalid file type. Please upload .xlsx.",
                "attachment_id": None,
            })

        content = fs.read()

        # Store XLSX as attachment so we can reuse it on "Create Users" button
        attachment = request.env["ir.attachment"].sudo().create({
            "name": fs.filename,
            "type": "binary",
            "datas": content.encode("base64") if isinstance(content, str) else __import__("base64").b64encode(content),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": "res.users",
        })

        try:
            wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active

            columns = self._normalize_columns([c.value for c in ws[1]])

            # read + count
            top10, count = [], 0
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if not row_cells or all((v is None or str(v).strip() == "") for v in row_cells):
                    continue
                count += 1
                if len(top10) < 10:
                    row_dict = {}
                    for idx, col in enumerate(columns):
                        val = row_cells[idx] if idx < len(row_cells) else ""
                        row_dict[col] = "" if val is None else val
                    top10.append(row_dict)

            upload_result = {"count": count, "columns": columns, "top10": top10}

            return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
                "active_menu": "sanctioned_posts",
                "upload_result": upload_result,
                "upload_error": None,
                "attachment_id": attachment.id,
            })

        except Exception as e:
            return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
                "active_menu": "sanctioned_posts",
                "upload_result": None,
                "upload_error": f"Failed to read XLSX: {e}",
                "attachment_id": None,
            })

    @http.route("/hrmis/users/import_create_all", type="http", auth="user", website=True, csrf=True, methods=["POST"])
    def import_create_all(self, **post):
        self._ensure_admin()

        attachment_id = int(post.get("attachment_id") or 0)
        if not attachment_id:
            raise UserError("Missing attachment_id. Please upload the XLSX again.")

        attachment = request.env["ir.attachment"].sudo().browse(attachment_id)
        if not attachment.exists():
            raise UserError("Uploaded file not found. Please upload again.")

        if openpyxl is None:
            raise UserError("openpyxl is not installed on the server.")

        content = base64.b64decode(attachment.datas or b"")
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        # Detect columns by header names
        header_vals = [c.value for c in ws[1]]
        columns = self._normalize_columns(header_vals)

        cnic_col = self._find_col(columns, ["cnic", "CNIC"])
        dob_col  = self._find_col(columns, ["dob", "DOB", "date of birth", "date_of_birth"])
        phone_col= self._find_col(columns, ["phone", "Phone", "mobile", "Mobile", "mobile_no", "mobile no"])

        name_col = self._find_col(columns, ["name", "Name", "employee_name", "employee name", "full_name", "full name"])

        if not cnic_col or not dob_col or not phone_col:
            # show user exactly what it saw
            raise UserError(
                "Missing required columns. Required: CNIC, DOB, Phone.\n"
                f"Detected columns: {', '.join(columns)}"
            )

        Users = request.env["res.users"].sudo()
        IrModelData = request.env["ir.model.data"].sudo()

        created = 0
        skipped = 0
        errors = []

        # Preload existing logins into a set for speed (optional but helps)
        # For huge DB, you can skip this and just search per row.
        # Here we do per row search to keep memory stable.

        batch = []
        batch_size = 500  # safe size
        rows_seen = 0

        def flush_batch():
            nonlocal created, batch
            if not batch:
                return
            new_users = Users.create(batch)
            created += len(new_users)

            # Create external IDs "use CNIC as ID" (safe equivalent)
            # xmlid: hrmis_user_profiles_updates.user_<cnic_digits>
            for u in new_users:
                cnic_digits = self._digits_only(u.login)
                xml_name = f"user_{cnic_digits}"
                # avoid duplicate xmlids
                if not IrModelData.search([("module", "=", "hrmis_user_profiles_updates"), ("name", "=", xml_name)], limit=1):
                    IrModelData.create({
                        "module": "hrmis_user_profiles_updates",
                        "name": xml_name,
                        "model": "res.users",
                        "res_id": u.id,
                    })

            request.env.cr.commit()
            batch = []

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            # skip empty rows
            if not row_cells or all((v is None or str(v).strip() == "") for v in row_cells):
                continue

            rows_seen += 1

            # Build a dict by column name
            row = {}
            for idx, col in enumerate(columns):
                row[col] = row_cells[idx] if idx < len(row_cells) else None

            cnic_raw = row.get(cnic_col)
            dob_raw = row.get(dob_col)
            phone_raw = row.get(phone_col)
            name_raw = row.get(name_col) if name_col else None

            cnic_login = self._get_cell_str(cnic_raw)
            if not cnic_login:
                skipped += 1
                if len(errors) < 50:
                    errors.append(f"Row {rows_seen+1}: Missing CNIC")
                continue

            # login should be cnic (as requested)
            login = cnic_login.strip()

            # skip duplicates
            if Users.search([("login", "=", login)], limit=1):
                skipped += 1
                continue

            dob_digits = self._parse_dob_to_digits(dob_raw)  # “without dashes/slashes”
            cnic_last4 = self._last4(cnic_login)
            phone_last4 = self._last4(phone_raw)

            # password format:
            # cnic last 4 digits + dob digits + phone last 4 digits
            password = f"{cnic_last4}{dob_digits}{phone_last4}"

            # name fallback
            name = self._get_cell_str(name_raw) or login

            batch.append({
                "name": name,
                "login": login,
                "password": password,          # real password field
                "hrmis_role": "employee",      # your custom field
                # optionally:
                # "email": row.get(email_col) ...
            })

            if len(batch) >= batch_size:
                flush_batch()

        flush_batch()

        # Show results on same page (reuse template)
        upload_result = {
            "count": rows_seen,
            "columns": columns,
            "top10": [],
            "created": created,
            "skipped": skipped,
            "errors": errors,
        }

        return request.render("hrmis_user_profiles_updates.hrmis_sanctioned_posts_upload_page", {
            "active_menu": "sanctioned_posts",
            "upload_result": upload_result,
            "upload_error": None if not errors else ("Some rows were skipped. See errors below."),
            "attachment_id": attachment.id,
        })