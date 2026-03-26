# -*- coding: utf-8 -*-
import json
import logging
from io import BytesIO
import uuid
from urllib.parse import quote

from odoo import http, fields
from odoo.http import request

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


QUEUE_NAME = "user_import"


def _norm(s):
    return (str(s or "")).strip().lower()


def _cell_str(v):
    if v is None:
        return ""
    # Prevent 10706671.0 from Excel numeric cells
    if isinstance(v, float) and v.is_integer():
        return str(int(v)).strip()
    return str(v).strip()


class HrmisUserConfigController(http.Controller):

    @http.route("/hrmis/user_config/recover/<int:job_id>", type="http", auth="user", website=True)
    def recover_user_import_job(self, job_id, **kw):

        q = request.env["hrmis.redis.queue"].sudo()
        Job = request.env["hrmis.user.import.job"].sudo()

        job = Job.browse(job_id)
        if not job.exists():
            return request.redirect("/hrmis/user_config?flash_error=Job not found")

        pending_q = f"{QUEUE_NAME}:{job.id}:pending"
        processing_q = f"{QUEUE_NAME}:{job.id}:processing"
        done_q = f"{QUEUE_NAME}:{job.id}:done"
        failed_q = f"{QUEUE_NAME}:{job.id}:failed"

        # --- Step 1: recover stuck processing items ---
        stuck_items = q.list_json(processing_q, 0, -1)
        recovered = 0

        for item in stuck_items:
            q.push_json(pending_q, item)
            q.lrem_json(processing_q, item, count=1)
            recovered += 1

        # --- Step 2: recalc counts from Redis ---
        pending = q.length(pending_q)
        processing = q.length(processing_q)
        done = q.length(done_q)
        failed = q.length(failed_q)

        job_vals = {
            "processed_count": done + failed,
            "created_count": done,
            "failed_count": failed,
        }

        if pending > 0 or processing > 0:
            job_vals["state"] = "queued"

        job.write(job_vals)

        # --- Step 3: finalize if nothing left ---
        if pending == 0 and processing == 0:
            worker = request.env["hrmis.user.queue.worker"].sudo()
            worker._finalize_job_with_report(job, failed_q)

        msg = f"Recovery executed. Recovered {recovered} stuck records."

        return request.redirect("/hrmis/user_config?flash_success=%s" % quote(msg))

    @http.route("/hrmis/user_config", type="http", auth="user", website=True, methods=["GET"], csrf=True)
    def user_config_page(self, **kw):
        Users = request.env["res.users"].sudo()

        today = fields.Date.today()
        start_dt = fields.Datetime.to_datetime(f"{today} 00:00:00")

        total_users = Users.search_count([])
        users_created_today = Users.search_count([("create_date", ">=", start_dt)])

        q = request.env["hrmis.redis.queue"].sudo()
        queue_len = q.length(QUEUE_NAME)
        latest_job = request.env["hrmis.user.import.job"].sudo().search([], limit=1, order="id desc")

        ctx = {
            "active_menu": "user_config",
            "metrics": {
                "total_users": total_users,
                "users_created_today": users_created_today,
                "queue_len": queue_len,
            },
            "flash_success": kw.get("flash_success"),
            "flash_error": kw.get("flash_error"),
            "last_queued_preview": kw.get("last_queued_preview"),
            "latest_job": latest_job,
        }
        return request.render("sanctioned_posts.hrmis_user_config_page", ctx)

    # @http.route("/hrmis/user_config/upload", type="http", auth="user", website=True, methods=["POST"], csrf=True)
    # def user_config_upload(self, **post):
    #     if openpyxl is None:
    #         return self._redirect_err("openpyxl not installed in container. Install it to parse XLSX.")

    #     f = request.httprequest.files.get("xlsx_file")
    #     if not f:
    #         return self._redirect_err("No file received.")

    #     sheet_name = (post.get("sheet_name") or "").strip()

    #     try:
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(f.read()),
    #             data_only=True,
    #             read_only=True,
    #         )

    #         if sheet_name:
    #             if sheet_name not in wb.sheetnames:
    #                 return self._redirect_err(
    #                     "Invalid sheet selected: %s. Available sheets: %s"
    #                     % (sheet_name, ", ".join(wb.sheetnames))
    #                 )
    #             ws = wb[sheet_name]
    #         else:
    #             ws = wb.active

    #     except Exception as e:
    #         return self._redirect_err(f"Failed to read XLSX: {e}")

    #     rows = ws.iter_rows(values_only=True)
    #     header_row = next(rows, None)

    #     if not header_row:
    #         return self._redirect_err("The uploaded file is empty.")

    #     headers = [_norm(h) for h in header_row]

    #     def find_col(*names):
    #         wanted = {_norm(n) for n in names}
    #         for idx, h in enumerate(headers):
    #             if h in wanted:
    #                 return idx
    #         return None

    #     col_pers_no = find_col("Pers.no.", "Pers no", "pers_no", "persno")
    #     col_name = find_col("Name at birth", "name at birth", "name")

    #     if col_pers_no is None or col_name is None:
    #         return self._redirect_err(
    #             "Missing required columns. Need: Pers.no., Name at birth."
    #         )

    #     q = request.env["hrmis.redis.queue"].sudo()
    #     Job = request.env["hrmis.user.import.job"].sudo()

    #     job = Job.create({
    #         "name": f"User Import - {fields.Datetime.now()}",
    #         "state": "queued",
    #         "queued_count": 0,
    #         "processed_count": 0,
    #         "created_count": 0,
    #         "failed_count": 0,
    #     })

    #     pending_q = f"{QUEUE_NAME}:{job.id}:pending"

    #     queued = 0
    #     preview_rows = []

    #     for excel_row_no, row in enumerate(rows, start=2):
    #         if not row or all(v is None or str(v).strip() == "" for v in row):
    #             continue

    #         pers_no = _cell_str(row[col_pers_no] if col_pers_no < len(row) else "")
    #         name = _cell_str(row[col_name] if col_name < len(row) else "")

    #         if not pers_no and not name:
    #             continue

    #         payload = {
    #             "job_id": job.id,
    #             "login": pers_no,
    #             "password": pers_no,
    #             "temp_password": pers_no,
    #             "name": name,
    #             "hrmis_role": "employee",
    #             "row": excel_row_no,
    #             "uploaded_by_uid": request.env.user.id,
    #             "uploaded_at": fields.Datetime.now().isoformat(),
    #             "queue_id": uuid.uuid4().hex,
    #         }

    #         ok = q.push_json(pending_q, payload)
    #         if ok:
    #             queued += 1
    #             if len(preview_rows) < 5:
    #                 preview_rows.append(payload)

    #     job.write({
    #         "queued_count": queued,
    #         "state": "queued",
    #     })

    #     today = fields.Date.today()
    #     start_dt = fields.Datetime.to_datetime(f"{today} 00:00:00")

    #     ctx = {
    #         "active_menu": "user_config",
    #         "metrics": {
    #             "total_users": request.env["res.users"].sudo().search_count([]),
    #             "users_created_today": request.env["res.users"].sudo().search_count([("create_date", ">=", start_dt)]),
    #             "queue_len": q.length(pending_q),
    #         },
    #         "flash_success": f"Queued {queued} users into Redis queue for Job #{job.id} (Sheet: {sheet_name or ws.title})",
    #         "flash_error": None,
    #         "last_queued_preview": json.dumps(preview_rows, indent=2, ensure_ascii=False),
    #         "latest_job": job,
    #         "job_pending_key": pending_q,
    #     }

    #     return request.render("sanctioned_posts.hrmis_user_config_page", ctx)

    @http.route("/hrmis/user_config/upload", type="http", auth="user", website=True, methods=["POST"], csrf=True)
    def user_config_upload(self, **post):
        if openpyxl is None:
            return self._redirect_err("openpyxl not installed in container. Install it to parse XLSX.")

        f = request.httprequest.files.get("xlsx_file")
        if not f:
            return self._redirect_err("No file received.")

        sheet_name = (post.get("sheet_name") or "").strip()

        try:
            wb = openpyxl.load_workbook(
                filename=BytesIO(f.read()),
                data_only=True,
                read_only=True,
            )

            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return self._redirect_err(
                        "Invalid sheet selected: %s. Available sheets: %s"
                        % (sheet_name, ", ".join(wb.sheetnames))
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active

        except Exception as e:
            return self._redirect_err(f"Failed to read XLSX: {e}")

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            return self._redirect_err("The uploaded file is empty.")

        headers = [_norm(h) for h in header_row]

        def find_col(*names):
            wanted = {_norm(n) for n in names}
            for idx, h in enumerate(headers):
                if h in wanted:
                    return idx
            return None

        def _cell_str(v):
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v)).strip()
            return str(v).strip()

        col_pers_no = find_col("Pers.no.", "Pers no", "pers_no", "persno")
        col_name = find_col("Personnel Number", "personnel number", "name")

        if col_pers_no is None or col_name is None:
            return self._redirect_err(
                "Missing required columns. Need: Pers.no., Personnel Number."
            )

        q = request.env["hrmis.redis.queue"].sudo()
        Job = request.env["hrmis.user.import.job"].sudo()
        Users = request.env["res.users"].sudo()

        # Step 1: collect rows from sheet first
        excel_rows = []
        all_pers_nos = set()

        for excel_row_no, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            pers_no = _cell_str(row[col_pers_no] if col_pers_no < len(row) else "")
            name = _cell_str(row[col_name] if col_name < len(row) else "")

            if not pers_no and not name:
                continue

            excel_rows.append({
                "row": excel_row_no,
                "login": pers_no,
                "password": pers_no,
                "temp_password": pers_no,
                "name": name,
                "hrmis_role": "employee",
            })

            if pers_no:
                all_pers_nos.add(pers_no)

        # Step 2: fetch existing users in ONE query
        existing_logins = set()
        if all_pers_nos:
            existing_users = Users.search([("login", "in", list(all_pers_nos))])
            existing_logins = set(existing_users.mapped("login"))

        # Optional: avoid duplicate Pers.no. inside same uploaded file too
        seen_in_file = set()

        job = Job.create({
            "name": f"User Import - {fields.Datetime.now()}",
            "state": "queued",
            "queued_count": 0,
            "processed_count": 0,
            "created_count": 0,
            "failed_count": 0,
        })

        pending_q = f"{QUEUE_NAME}:{job.id}:pending"

        queued = 0
        skipped_existing = 0
        skipped_duplicate_in_file = 0
        preview_rows = []

        for row_data in excel_rows:
            login = row_data.get("login", "")

            if not login:
                continue

            # Skip if already exists in DB
            if login in existing_logins:
                skipped_existing += 1
                continue

            # Skip duplicate Pers.no. repeated within same Excel file
            if login in seen_in_file:
                skipped_duplicate_in_file += 1
                continue
            seen_in_file.add(login)

            payload = {
                "job_id": job.id,
                "login": row_data["login"],
                "password": row_data["password"],
                "temp_password": row_data["temp_password"],
                "name": row_data["name"],
                "hrmis_role": "employee",
                "row": row_data["row"],
                "uploaded_by_uid": request.env.user.id,
                "uploaded_at": fields.Datetime.now().isoformat(),
                "queue_id": uuid.uuid4().hex,
            }

            ok = q.push_json(pending_q, payload)
            if ok:
                queued += 1
                if len(preview_rows) < 5:
                    preview_rows.append(payload)

        job.write({
            "queued_count": queued,
            "state": "queued",
        })

        today = fields.Date.today()
        start_dt = fields.Datetime.to_datetime(f"{today} 00:00:00")

        msg = (
            f"Queued {queued} users into Redis queue for Job #{job.id} "
            f"(Sheet: {sheet_name or ws.title}). "
            f"Skipped existing users: {skipped_existing}. "
            f"Skipped duplicate Pers.no. in file: {skipped_duplicate_in_file}."
        )

        ctx = {
            "active_menu": "user_config",
            "metrics": {
                "total_users": request.env["res.users"].sudo().search_count([]),
                "users_created_today": request.env["res.users"].sudo().search_count([("create_date", ">=", start_dt)]),
                "queue_len": q.length(pending_q),
            },
            "flash_success": msg,
            "flash_error": None,
            "last_queued_preview": json.dumps(preview_rows, indent=2, ensure_ascii=False),
            "latest_job": job,
            "job_pending_key": pending_q,
        }

        return request.render("sanctioned_posts.hrmis_user_config_page", ctx)

    def _redirect_err(self, msg: str):
        return request.redirect("/hrmis/user_config?flash_error=%s" % quote(msg or "", safe=""))

    @http.route("/hrmis/user_config/queue", type="http", auth="user", website=True, methods=["GET"], csrf=True)
    def user_queue_page(self, **kw):
        q = request.env["hrmis.redis.queue"].sudo()
        Job = request.env["hrmis.user.import.job"].sudo()

        latest_job = Job.search([], limit=1, order="id desc")

        pending = processing = done = failed = 0
        pending_list = []
        failed_list = []
        done_list = []

        pending_q = processing_q = done_q = failed_q = None

        if latest_job:
            pending_q = f"{QUEUE_NAME}:{latest_job.id}:pending"
            processing_q = f"{QUEUE_NAME}:{latest_job.id}:processing"
            done_q = f"{QUEUE_NAME}:{latest_job.id}:done"
            failed_q = f"{QUEUE_NAME}:{latest_job.id}:failed"

            pending = q.length(pending_q)
            processing = q.length(processing_q)
            done = q.length(done_q)
            failed = q.length(failed_q)

            pending_list = q.list_json(pending_q, 0, 9)
            failed_list = q.list_json(failed_q, 0, 9)
            done_list = q.list_json(done_q, 0, 9)

        ctx = {
            "active_menu": "user_config",
            "latest_job": latest_job,
            "pending": pending,
            "processing": processing,
            "done": done,
            "failed": failed,
            "pending_list": pending_list,
            "failed_list": failed_list,
            "done_list": done_list,
            "pending_q": pending_q,
            "processing_q": processing_q,
            "done_q": done_q,
            "failed_q": failed_q,
        }
        return request.render("sanctioned_posts.hrmis_user_queue_page", ctx)

    @http.route("/hrmis/user_config/process", type="http", auth="user", website=True, methods=["GET"], csrf=False)
    def user_config_process(self, **kw):
        return request.redirect("/hrmis/user_config/queue")


    @http.route("/hrmis/user_config/rename_users", type="http", auth="user", website=True, methods=["POST"], csrf=True)
    def rename_users_from_xlsx(self, **post):
        if openpyxl is None:
            return self._redirect_err("openpyxl not installed in container. Install it to parse XLSX.")

        f = request.httprequest.files.get("rename_xlsx_file")
        if not f:
            return self._redirect_err("No file received.")

        try:
            wb = openpyxl.load_workbook(
                filename=BytesIO(f.read()),
                data_only=True,
                read_only=True,
            )
            ws = wb.active
        except Exception as e:
            return self._redirect_err(f"Failed to read XLSX: {e}")

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            return self._redirect_err("The uploaded file is empty.")

        headers = [_norm(h) for h in header_row]

        def find_col(*names):
            wanted = {_norm(n) for n in names}
            for idx, h in enumerate(headers):
                if h in wanted:
                    return idx
            return None

        col_pers_no = find_col("Pers.no.", "Pers no", "pers_no", "persno")
        col_personnel_number = find_col("Personnel Number", "personnel number", "personnel_number")

        if col_pers_no is None or col_personnel_number is None:
            return self._redirect_err(
                "Missing required columns. Need: Pers.no., Personnel Number."
            )

        Users = request.env["res.users"].sudo()

        # Collect excel rows first
        excel_data = []
        all_pers_nos = set()

        for excel_row_no, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            pers_no = _cell_str(row[col_pers_no] if col_pers_no < len(row) else "")
            personnel_number = _cell_str(row[col_personnel_number] if col_personnel_number < len(row) else "")

            if not pers_no:
                continue

            excel_data.append({
                "row": excel_row_no,
                "pers_no": pers_no,
                "personnel_number": personnel_number,
            })
            all_pers_nos.add(pers_no)

        if not excel_data:
            return self._redirect_err("No valid data rows found in the uploaded file.")

        # Fetch all matching users in one query
        users = Users.search([("login", "in", list(all_pers_nos))])
        users_by_login = {str(u.login).strip(): u for u in users}

        updated_count = 0
        skipped_not_found = 0
        skipped_blank_name = 0
        preview_rows = []

        for item in excel_data:
            user = users_by_login.get(item["pers_no"])
            new_name = item["personnel_number"]

            if not user:
                skipped_not_found += 1
                continue

            if not new_name:
                skipped_blank_name += 1
                continue

            old_name = user.name
            user.write({
                "name": new_name,
            })
            updated_count += 1

            if len(preview_rows) < 10:
                preview_rows.append({
                    "row": item["row"],
                    "login": item["pers_no"],
                    "old_name": old_name,
                    "new_name": new_name,
                })

        msg = (
            f"Rename completed. Updated users: {updated_count}. "
            f"Users not found: {skipped_not_found}. "
            f"Skipped blank Personnel Number: {skipped_blank_name}."
        )

        today = fields.Date.today()
        start_dt = fields.Datetime.to_datetime(f"{today} 00:00:00")
        q = request.env["hrmis.redis.queue"].sudo()
        latest_job = request.env["hrmis.user.import.job"].sudo().search([], limit=1, order="id desc")

        ctx = {
            "active_menu": "user_config",
            "metrics": {
                "total_users": Users.search_count([]),
                "users_created_today": Users.search_count([("create_date", ">=", start_dt)]),
                "queue_len": q.length(QUEUE_NAME),
            },
            "flash_success": msg,
            "flash_error": None,
            "last_queued_preview": json.dumps(preview_rows, indent=2, ensure_ascii=False),
            "latest_job": latest_job,
        }
        return request.render("sanctioned_posts.hrmis_user_config_page", ctx)