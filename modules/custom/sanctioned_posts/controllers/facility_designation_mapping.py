# -*- coding: utf-8 -*-
import logging
from io import BytesIO
from urllib.parse import quote

from odoo import http
from odoo.http import request

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except Exception:
    openpyxl = None


def _norm(value):
    return (str(value or "").strip().upper())


def _cell_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


class HrmisSanctionedPostsUploadController(http.Controller):

    @http.route(
        "/hrmis/sanctioned_posts/upload_designations",
        type="http",
        auth="user",
        website=True,
        methods=["POST"],
        csrf=True,
    )
    def upload_designations_xlsx(self, **post):
        if not request.env.user.has_group("base.group_system"):
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote("You are not allowed to perform this action.", safe="")
            )

        if openpyxl is None:
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote("openpyxl not installed in container.", safe="")
            )

        uploaded_file = request.httprequest.files.get("designations_xlsx_file")
        if not uploaded_file:
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote("No file received.", safe="")
            )

        sheet_name = (post.get("sheet_name") or "").strip()

        try:
            workbook = openpyxl.load_workbook(
                filename=BytesIO(uploaded_file.read()),
                data_only=True,
                read_only=True,
            )

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return request.redirect(
                        "/hrmis/sanctioned_posts/upload?flash_error=%s"
                        % quote(
                            "Invalid sheet selected: %s. Available sheets: %s"
                            % (sheet_name, ", ".join(workbook.sheetnames)),
                            safe="",
                        )
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

        except Exception as e:
            _logger.exception("Failed to read XLSX")
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote("Failed to read XLSX: %s" % e, safe="")
            )

        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote("The uploaded file is empty.", safe="")
            )

        headers = [_norm(h) for h in header_row]

        def find_col(*names):
            wanted = {_norm(name) for name in names}
            for idx, header in enumerate(headers):
                if header in wanted:
                    return idx
            return None

        col_designation = find_col(
            "DESIGNATION",
            "DESIGNATIONS",
            "DESIGNATION NAME",
            "NAME",
        )
        col_level_of_care = find_col(
            "LEVEL OF CARE",
            "LEVEL_OF_CARE",
            "LEVEL",
        )
        col_facility_name = find_col(
            "FACILITY NAME",
            "FACILITY_NAME",
            "FACILITY",
        )

        if col_designation is None or col_level_of_care is None:
            return request.redirect(
                "/hrmis/sanctioned_posts/upload?flash_error=%s"
                % quote(
                    "Missing required columns. Need at least: Designation and Level of Care. Facility Name is optional.",
                    safe="",
                )
            )

        LevelCareDesignation = request.env["hrmis.level.care.designation"].sudo()

        created = 0
        skipped = 0
        errors = []

        for excel_row_no, row in enumerate(rows, start=2):
            try:
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue

                designation_name = _cell_str(
                    row[col_designation] if col_designation < len(row) else ""
                )
                level_of_care = _cell_str(
                    row[col_level_of_care] if col_level_of_care < len(row) else ""
                )

                facility_name = ""
                if col_facility_name is not None and col_facility_name < len(row):
                    facility_name = _cell_str(row[col_facility_name])

                if not designation_name:
                    skipped += 1
                    if len(errors) < 20:
                        errors.append("Row %s: Designation is empty." % excel_row_no)
                    continue

                vals = {
                    "name": designation_name,
                    "level_of_care": level_of_care or False,
                    "designation_group_id": 1,
                    "total_sanctioned_posts": 1,
                    "post_BPS": 0,
                    "facility_id": 1,
                    "active": True,
                    "old_value": False,
                    "is_temp": False,
                }

                if facility_name:
                    vals["facility_name"] = facility_name

                LevelCareDesignation.create(vals)
                created += 1

            except Exception as e:
                skipped += 1
                _logger.exception("Error on Excel row %s", excel_row_no)
                if len(errors) < 20:
                    errors.append("Row %s: %s" % (excel_row_no, str(e)))

        msg = (
            "Designation upload completed. "
            "Created: %s. Skipped: %s. Sheet: %s"
            % (created, skipped, sheet_name or worksheet.title)
        )

        if errors:
            msg += " Sample errors: " + " | ".join(errors[:5])

        return request.redirect(
            "/hrmis/sanctioned_posts/upload?flash_success=%s"
            % quote(msg, safe="")
        )