# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import logging
from datetime import datetime

from odoo import http, fields
from odoo.http import request
from odoo.exceptions import AccessError

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

_logger = logging.getLogger(__name__)


class HrmisEmployeeExportController(http.Controller):

    @http.route(
        ["/hrmis/employees/download"],
        type="http",
        auth="user",
        methods=["GET"],
        csrf=False,
    )
    def hrmis_employees_download(self, ids="all", **kw):
        """
        Returns XLSX. ids=all OR ids=1,2,3
        """
        user = request.env.user
        _logger.info("[EMP_EXPORT] Download requested. ids=%s user=%s(%s)", ids, user.name, user.id)

        # Access guard: limit to HR users/managers or system
        if not (
            user.has_group("hr.group_hr_user")
            or user.has_group("hr.group_hr_manager")
            or user.has_group("base.group_system")
        ):
            _logger.warning("[EMP_EXPORT] Access denied for user=%s(%s)", user.name, user.id)
            raise AccessError("You are not allowed to export employee data.")

        Employee = request.env["hr.employee"].with_context(active_test=False)

        # Resolve employees
        if ids == "all":
            employees = Employee.search([])
        else:
            try:
                emp_ids = [int(x) for x in (ids or "").split(",") if x.strip()]
            except Exception:
                _logger.exception("[EMP_EXPORT] Invalid ids param: %s", ids)
                emp_ids = []
            employees = Employee.browse(emp_ids).exists()

        _logger.info("[EMP_EXPORT] Employees resolved: count=%s", len(employees))

        # Prefetch / models
        ProfileReq = request.env["hrmis.employee.profile.request"]
        Posting = request.env["hrmis.posting.history"]
        Promotion = request.env["hrmis.promotion.history"]
        Qualif = request.env["hrmis.qualification.history"]
        Leave = request.env["hrmis.leave.history"]

        # Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "Employee Name",
            "Work Email",
            "User Login",
            "Employee ID / Service #",
            "CNIC",
            "Father Name",
            "DOB",
            "Gender",
            "Cadre",
            "Designation",
            "BPS",
            "District",
            "Facility",
            "Contact Info",
            "Domicile",
            "Qualification",
            "Qualification Date",
            "Year of Qualification",
            "Last Promotion Date",
            "Promotion BPS From",
            "Promotion BPS To",
            "Current Posting District",
            "Current Posting Facility",
            "Current Posting Designation",
            "Current Posting Start",
            "Last Posting Start",
            "Last Posting End",
            "Leaves Count",
            "Last Leave Start",
            "Last Leave End",
            "Latest Profile Request State",
            "Latest Profile Request Submitted/Approved By",
            "Latest Profile Request Date",
        ]

        ws.append(headers)
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _fmt_date(d):
            return d.strftime("%Y-%m-%d") if d else ""

        row_count = 1

        for emp in employees:
            # Related user/login
            user_login = emp.user_id.login if emp.user_id else ""

            # Latest profile request (prefer approved, else newest)
            latest_req = ProfileReq.search(
                [("employee_id", "=", emp.id)],
                order="id desc",
                limit=1,
            )
            latest_req_state = latest_req.state if latest_req else ""
            latest_req_actor = ""
            latest_req_date = ""
            if latest_req:
                # best-effort: approved_by if set, else approver_id, else user_id
                if getattr(latest_req, "approved_by", False):
                    latest_req_actor = latest_req.approved_by.name or ""
                elif latest_req.approver_id:
                    latest_req_actor = latest_req.approver_id.name or ""
                elif latest_req.user_id:
                    latest_req_actor = latest_req.user_id.name or ""
                # mail.thread records have write_date/create_date; use write_date
                latest_req_date = _fmt_date(latest_req.write_date.date() if latest_req.write_date else None)

            # Posting: current + latest
            current_posting = Posting.search(
                [("employee_id", "=", emp.id), ("is_current", "=", True)],
                order="start_date desc, id desc",
                limit=1,
            )
            last_posting = Posting.search(
                [("employee_id", "=", emp.id)],
                order="start_date desc, id desc",
                limit=1,
            )

            # Promotion: latest
            last_promo = Promotion.search(
                [("employee_id", "=", emp.id)],
                order="promotion_date desc, id desc",
                limit=1,
            )

            # Qualification: latest
            last_qual = Qualif.search(
                [("employee_id", "=", emp.id)],
                order="start_date desc, id desc",
                limit=1,
            )

            # Leaves: count + last
            leaves_count = Leave.search_count([("employee_id", "=", emp.id)])
            last_leave = Leave.search(
                [("employee_id", "=", emp.id)],
                order="start_date desc, id desc",
                limit=1,
            )

            # Safely resolve names (Many2one)
            def _m2o_name(rec):
                return rec.name if rec else ""

            # For posting designation_id is Many2one (hrmis.designation)
            def _posting_designation_name(post):
                return post.designation_id.name if post and post.designation_id else ""

            ws.append([
                emp.name or "",
                emp.work_email or "",
                user_login,
                getattr(emp, "hrmis_employee_id", "") or "",
                getattr(emp, "hrmis_cnic", "") or "",
                getattr(emp, "hrmis_father_name", "") or "",
                _fmt_date(emp.birthday),
                emp.gender or "",
                _m2o_name(getattr(emp, "hrmis_cadre", False)),
                _m2o_name(getattr(emp, "hrmis_designation", False)) if hasattr(emp, "hrmis_designation") else "",
                getattr(emp, "hrmis_bps", "") or "",
                _m2o_name(getattr(emp, "district_id", False)),
                _m2o_name(getattr(emp, "facility_id", False)),
                getattr(emp, "hrmis_contact_info", "") or "",
                getattr(emp, "hrmis_domicile", "") or "",
                getattr(emp, "qualification", "") or "",
                _fmt_date(getattr(emp, "qualification_date", False)),
                _fmt_date(getattr(emp, "year_qualification", False)),
                _fmt_date(last_promo.promotion_date) if last_promo else "",
                last_promo.bps_from if last_promo else "",
                last_promo.bps_to if last_promo else "",
                _m2o_name(current_posting.district_id) if current_posting else "",
                _m2o_name(current_posting.facility_id) if current_posting else "",
                _posting_designation_name(current_posting) if current_posting else "",
                _fmt_date(current_posting.start_date) if current_posting else "",
                _fmt_date(last_posting.start_date) if last_posting else "",
                _fmt_date(last_posting.end_date) if last_posting else "",
                leaves_count,
                _fmt_date(last_leave.start_date) if last_leave else "",
                _fmt_date(last_leave.end_date) if last_leave else "",
                latest_req_state,
                latest_req_actor,
                latest_req_date,
            ])
            row_count += 1

        _logger.info("[EMP_EXPORT] XLSX rows written (including header): %s", row_count)

        # Light formatting
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Autosize columns (simple heuristic)
        for col_idx in range(1, len(headers) + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(
                max(12, len(headers[col_idx - 1]) + 2),
                40
            )

        # Output
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"employee_info_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        _logger.info("[EMP_EXPORT] Sending file: %s", filename)

        return request.make_response(
            bio.getvalue(),
            headers=[
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
            ],
        )
